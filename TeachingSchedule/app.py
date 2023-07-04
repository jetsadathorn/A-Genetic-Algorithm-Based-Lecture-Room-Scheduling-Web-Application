from flask import Flask, render_template, request, redirect, url_for, make_response
from openpyxl.utils import get_column_letter
from io import BytesIO
from openpyxl import Workbook
import os
import pandas as pd
from flask_cors import CORS
import random as rnd
import mysql.connector
import prettytable as prettytable
from enum import Enum
POPULATION_SIZE = 9  # จำนวนประชากรในแอลกอริทึม
NUMB_OF_ELITE_SCHEDULES = 1  # จำนวนตารางสอนที่เลือกไว้ค่า fitness ที่สูงที่สุด
# จำนวนตัวอย่างที่เลือกในการทัวร์นาเมนต์ในการเลือกพ่อแม่
TOURNAMENT_SELECTION_SIZE = 3
MUTATION_RATE = 0.1  # อัตราการระบายพันธุ์ในการแปลงตารางสอน
VERBOSE_FLAG = False


app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False
CORS(app)

# Upload folder
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


class DBMgr:
    def __init__(self):
        self._conn = mysql.connector.connect(
            host="localhost", user="root", password="", db="time_tabling", buffered=True)
        self._c = self._conn.cursor()
        self._rooms = self._select_rooms()
        self._meetingTimes = self._select_meeting_times()
        self._instructors = self._select_instructors()
        self._courses = self._select_courses()
        self._depts = self._select_depts()
        self._numberOfClasses = 0
        for i in range(0, len(self._depts)):
            self._numberOfClasses += len(self._depts[i].get_courses())

    def _select_rooms(self):
        self._c.execute("SELECT * FROM room")
        rooms = self._c.fetchall()
        returnRooms = []
        for i in range(0, len(rooms)):
            returnRooms.append(Room(rooms[i][0], rooms[i][1]))
        return returnRooms

    def _select_meeting_times(self):
        self._c.execute("SELECT * FROM meeting_time")
        meetingTimes = self._c.fetchall()
        returnMeetingTimes = []
        for i in range(0, len(meetingTimes)):
            returnMeetingTimes.append(MeetingTime(
                meetingTimes[i][0], meetingTimes[i][1], meetingTimes[i][2]))
        return returnMeetingTimes

    def _select_instructors(self):
        self._c.execute("SELECT * FROM instructor")
        instructors = self._c.fetchall()
        returnInstructors = []
        for i in range(0, len(instructors)):
            returnInstructors.append(Instructor(
                instructors[i][0], instructors[i][1], self._select_instructor_availability(instructors[i][0])))
        return returnInstructors

    def _select_instructor_availability(self, instructor):
        self._c.execute(
            "SELECT * from instructor_availability where instructor_id = '" + instructor + "'")
        instructorMTsRS = self._c.fetchall()
        instructorMTs = []
        for i in range(0, len(instructorMTsRS)):
            instructorMTs.append(instructorMTsRS[i][1])
        instructorAvailability = list()
        for i in range(0, len(self._meetingTimes)):
            if self._meetingTimes[i].get_id() in instructorMTs:
                instructorAvailability.append(self._meetingTimes[i])
        return instructorAvailability

    def _select_courses(self):
        self._c.execute("SELECT * FROM course")
        courses = self._c.fetchall()
        returnCourses = []
        for i in range(0, len(courses)):
            returnCourses.append(
                Course(courses[i][0], courses[i][1], self._select_course_instructors(courses[i][0]), courses[i][2]))
        return returnCourses

    def _select_depts(self):
        self._c.execute("SELECT * FROM dept")
        depts = self._c.fetchall()
        returnDepts = []
        for i in range(0, len(depts)):
            returnDepts.append(Department(
                depts[i][0], self._select_dept_courses(depts[i][0])))
        return returnDepts

    def _select_course_instructors(self, courseNumber):
        self._c.execute(
            "SELECT * FROM course_instructor where course_number = '" + courseNumber + "'")
        dbInstructorNumbers = self._c.fetchall()
        instructorNumbers = []
        for i in range(0, len(dbInstructorNumbers)):
            instructorNumbers.append(dbInstructorNumbers[i][1])
        returnValue = []
        for i in range(0, len(self._instructors)):
            if self._instructors[i].get_id() in instructorNumbers:
                returnValue.append(self._instructors[i])
        return returnValue

    def _select_dept_courses(self, deptName):
        self._c.execute(
            "SELECT * FROM dept_course where name = '" + deptName + "'")
        dbCourseNumbers = self._c.fetchall()
        courseNumbers = []
        for i in range(0, len(dbCourseNumbers)):
            courseNumbers.append(dbCourseNumbers[i][1])
        returnValue = []
        for i in range(0, len(self._courses)):
            if self._courses[i].get_number() in courseNumbers:
                returnValue.append(self._courses[i])
        return returnValue

    def get_rooms(self): return self._rooms
    def get_instructors(self): return self._instructors
    def get_courses(self): return self._courses
    def get_depts(self): return self._depts
    def get_meetingTimes(self): return self._meetingTimes
    def get_numberOfClasses(self): return self._numberOfClasses


class Schedule:
    def __init__(self):
        self._data = dbMgr
        self._classes = []
        self._conflicts = []
        self._fitness = -1
        self._classNumb = 0
        self._isFitnessChanged = True

    def get_classes(self):
        self._isFitnessChanged = True
        return self._classes

    def get_conflicts(self): return self._conflicts

    def get_fitness(self):
        if (self._isFitnessChanged == True):
            self._fitness = self.calculate_fitness()
            self._isFitnessChanged = False
        return self._fitness

    # generate ตัวใหม่มา
    def initialize(self):
        depts = self._data.get_depts()
        for i in range(0, len(depts)):
            courses = depts[i].get_courses()
            for j in range(0, len(courses)):
                newClass = Class(self._classNumb, depts[i], courses[j])
                self._classNumb += 1
                # สุ่ม เวลา ห้อง
                newClass.set_meetingTime(dbMgr.get_meetingTimes(
                )[rnd.randrange(0, len(dbMgr.get_meetingTimes()))])
                newClass.set_room(dbMgr.get_rooms()[
                                  rnd.randrange(0, len(dbMgr.get_rooms()))])
                newClass.set_instructor(courses[j].get_instructors(
                )[rnd.randrange(0, len(courses[j].get_instructors()))])
                self._classes.append(newClass)
        return self

    def calculate_fitness(self):
        dbMgr._c.execute("SELECT * FROM multi_sec")
        multi_sec = dbMgr._c.fetchall()
        result_dict = {key: value for key, value in multi_sec}

        dbMgr._c.execute("SELECT * FROM exp_dept")
        exp_dept = dbMgr._c.fetchall()
        dbMgr._c.execute("SELECT * FROM parallel")
        parallel = dbMgr._c.fetchall()

        self._numbOfConflicts = 0
        self._conflicts = []
        classes = self.get_classes()
        groups = {}

        for i in range(0, len(classes)):
            seatingCapacityConflict = []
            meetingTimeConflict = []
            seatingCapacityConflict.append(classes[i])
            # ความจุวิชา > ความจุห้อง
            if classes[i].get_room().get_capacity() < classes[i].get_course().get_maxNumbOfStudents():
                self._conflicts.append(
                    Conflict(Conflict.ConflictType.จำนวนนักศึกษาเกิน, seatingCapacityConflict))
                self._numbOfConflicts += 1

            # ถ้าเวลาเรียนของชั้นปีนั้นๆ ทับกับวันยกเว้นการสอน
            if (classes[i].get_meetingTime().get_id() in [x[1] for x in exp_dept if x[0] == classes[i].get_dept().get_name()]):
                conflictBetweenClasses = []
                conflictBetweenClasses.append(classes[i])
                self._conflicts.append(Conflict(
                    Conflict.ConflictType.ยกเว้นการสอนของชั้นปี, conflictBetweenClasses))
                self._numbOfConflicts += 1

            if (classes[i].get_meetingTime() in classes[i].get_instructor().get_availability()):
                conflictBetweenClasses = []
                conflictBetweenClasses.append(classes[i])
                self._conflicts.append(Conflict(
                    Conflict.ConflictType.INSTRUCTOR_AVAILABILITY, conflictBetweenClasses))
                self._numbOfConflicts += 1

            # หารหัสที่ต้องการ
            if classes[i].get_course().get_number() in [x[0] for x in multi_sec]:
                for x in classes:
                    if x.get_course().get_number() == result_dict[classes[i].get_course().get_number()]:
                        targetClass = x
                        break
                # ถ้าเวลาของ 2 วิชาไม่เหมือนกัน
                if classes[i].get_meetingTime().get_id() != targetClass.get_meetingTime().get_id():
                    self._conflicts.append(
                        Conflict(Conflict.ConflictType.ไม่ไขว้, seatingCapacityConflict))
                    self._numbOfConflicts += 1

            # สอนขนาน
            elif (classes[i].get_course().get_number() in [x[0] for x in parallel]):
                for q in range(0, len(classes)):
                    if f'{classes[i].get_course().get_number()}A' == classes[q].get_course().get_number():
                        if classes[i].get_meetingTime().get_time() != classes[q].get_meetingTime().get_time():
                            conflictBetweenClasses = []
                            conflictBetweenClasses.append(classes[i])
                            self._conflicts.append(Conflict(
                                Conflict.ConflictType.สอนขนาน, conflictBetweenClasses))
                            self._numbOfConflicts += 1
                            # break
            else:
                # เวลาสอนในหนึ่งชั้นปี เหมือนกัน
                dept = classes[i].get_dept()
                if dept not in groups:
                    groups[dept] = set()

                meeting_time = classes[i].get_meetingTime()
                if meeting_time in groups[dept]:
                    meetingTimeConflict.append(classes[i])
                    self._conflicts.append(
                        Conflict(Conflict.ConflictType.ปีมีวันซ้ำกัน, meetingTimeConflict))
                    self._numbOfConflicts += 1
                else:
                    groups[dept].add(meeting_time)

            for j in range(0, len(classes)):
                if (j >= i):
                    # เช็คต่อว่า ถ้าเวลาสอนเหมือนกัน
                    if (classes[i].get_meetingTime() == classes[j].get_meetingTime()
                            and classes[i].get_id() != classes[j].get_id()):
                        # แล้วห้องเดียวกันด้วย
                        if (classes[i].get_room() == classes[j].get_room()):
                            roomBookingConflict = [classes[i], classes[j]]
                            self._conflicts.append(
                                Conflict(Conflict.ConflictType.ROOM_BOOKING, roomBookingConflict))
                            self._numbOfConflicts += 1

                        # แล้วถ้าเวลาสอนเหมือนกัน แล้วอาจารย์คนเดียวกัน
                        if (classes[i].get_meetingTime() == classes[j].get_meetingTime() and classes[i].get_instructor() == classes[j].get_instructor()):
                            instructorBookingConflict = [
                                classes[i], classes[j]]
                            self._conflicts.append(
                                Conflict(Conflict.ConflictType.INSTRUCTOR_BOOKING, instructorBookingConflict))
                            self._numbOfConflicts += 1

        # Update fitness value
        self._fitness = 1 / (1.0*(self._numbOfConflicts) + 1)
        return self._fitness

    def __str__(self):
        returnValue = ""
        for i in range(0, len(self._classes)-1):
            returnValue += str(self._classes[i]) + "\n"
        returnValue += str(self._classes[len(self._classes)-1])
        return returnValue


# สุ่ม เวลา ห้อง อาจารย์ ทุกclass
class Population:
    def __init__(self, size):
        self._size = size
        self._data = dbMgr
        self._schedules = []
        for i in range(0, size):
            self._schedules.append(Schedule().initialize())

    def get_schedules(self): return self._schedules


class GeneticAlgorithm:
    # ส่งข้อมูลกลับไป
    def evolve(self, population):
        return self._mutate_population(self._selection(population))

    # Selection เลือกพ่อแม่
    def _selection(self, pop):
        # สร้างมาเปล่าๆ
        crossover_pop = Population(0)
        # ตัวค่า fitness ที่ดีที่สุดออกมา
        for i in range(NUMB_OF_ELITE_SCHEDULES):  # 1 ตัว
            crossover_pop.get_schedules().append(pop.get_schedules()[i])

            # print("ตัวค่า fitness ที่ดีที่สุด: ", i+1)#
            # DisplayMgr.display_schedule_as_table(
            #     crossover_pop.get_schedules()[i])#
            # print("-"*300)#

        i = NUMB_OF_ELITE_SCHEDULES  # = 1
        while i < POPULATION_SIZE:  # เลือกออกมาให้ครบทั้ง 9 ตัว
            # เลือก schedule ที่ดีที่สุดมาเก็บ (ตำแหน่งที่0)

            # print("คู่ที่ ", i)#
            # print("schedule1")#

            schedule1 = self._select_tournament_population(pop).get_schedules()[
                0]

            # DisplayMgr.display_schedule_as_table(schedule1)#
            # print()#
            # print("schedule2")#

            schedule2 = self._select_tournament_population(pop).get_schedules()[
                0]

            # DisplayMgr.display_schedule_as_table(schedule2)#
            # print("\n")#

            # จะมี schedule1,2 ทั้งหมด 9 คู่
            # นำ crossoverSchedule ไปต่อกับ schedule ที่ค่า fitness ดีที่สุด จะได้ทั้งหมด 9 ตัว
            crossover_pop.get_schedules().append(
                self._crossover_schedule(schedule1, schedule2))
            i += 1

        # print("ผลลัพท์ที่ได้จากการ Crossover")#
        # for i in range(0, len(crossover_pop.get_schedules())):#
        #     DisplayMgr.display_schedule_as_table(#
        #         crossover_pop.get_schedules()[i])#
        # print("-"*300)#

        # 8 ตัวที่เหลือ จะถูกเก็บไว้ใน crossover
        return crossover_pop

    def _select_tournament_population(self, pop):
        # สร้าง tournament_pop มาเปล่าๆ
        tournament_pop = Population(0)
        i = 0
        # ทำทั้งหมด 3 รอบ
        while i < TOURNAMENT_SELECTION_SIZE:  # < 3
            # เลือกมา 3 schedule จากทั้งหมด 9 ตัว
            random_index = rnd.randrange(0, POPULATION_SIZE)
            tournament_pop.get_schedules().append(
                pop.get_schedules()[random_index])
            # print("schedules ตัวที่: ", random_index+1)#
            i += 1
        # เอา 3 ตัวนี้ไปคิดค่า fitness แล้วเรียงจากมากไปน้อย
        tournament_pop.get_schedules().sort(key=lambda x: x.get_fitness(), reverse=True)
        return tournament_pop
        # crossover

    def _crossover_schedule(self, schedule1, schedule2):
        # สร้าง class มาใหม่แบบสุ่ม
        crossoverSchedule = Schedule().initialize()
        # สลับ schedule1 และ schedule2
        random_index = rnd.randint(
            0, len(crossoverSchedule.get_classes()) - 1)
        for i in range(0, random_index):
            crossoverSchedule.get_classes()[i] = schedule1.get_classes()[i]
        for i in range(random_index, len(crossoverSchedule.get_classes())):
            crossoverSchedule.get_classes()[i] = schedule2.get_classes()[i]

        # print("cross")#
        # print("ตำแหน่งยีนส์ที่สุ่มได้ ", random_index)#
        # DisplayMgr.display_schedule_as_table(crossoverSchedule)#
        # print("-"*300)#

        # เก็บค่าไว้ใน crossoverSchedule
        return crossoverSchedule

    # matate ที่เหลือ ที่ไม่ใช้ค่า fitness ที่ดีที่สุด
    def _mutate_population(self, population):
        for i in range(NUMB_OF_ELITE_SCHEDULES, POPULATION_SIZE):
            # print(f"schedule {i} :", end=" ")#
            self._mutate_schedule(population.get_schedules()[i])
        return population

    # mutate ค่าข้างใน
    def _mutate_schedule(self, mutateSchedule):
        # สร้าง schedule ขึ้นมาใหม่ 1 ตัว แบบสุ่ม
        schedule = Schedule().initialize()
        # print("schedule Mutate:")#
        # สุ่ม class ข้างใน schedule ทั้งหมด มีโอกาส 10 % ที่จะกลายพันธุ์
        count = 0
        for i in range(0, len(mutateSchedule.get_classes())):
            if (MUTATION_RATE > rnd.random()):  # 0.1,10% random ค่า 0-1
                count += 1
                mutateSchedule.get_classes()[i] = schedule.get_classes()[i]
        # DisplayMgr.display_schedule_as_table(mutateSchedule)#
        # print(f" กลายพันธุ์ {count} ตัว")#
        return mutateSchedule


class Course:
    def __init__(self, number, name, instructors, maxNumbOfStudents):
        self._number = number
        self._name = name
        self._maxNumbOfStudents = maxNumbOfStudents
        self._instructors = instructors

    def get_number(self): return self._number
    def get_name(self): return self._name
    def get_instructors(self): return self._instructors
    def get_maxNumbOfStudents(self): return self._maxNumbOfStudents
    def __str__(self): return self._name


class Instructor:
    def __init__(self, id, name, availability):
        self._id = id
        self._name = name
        self._availability = availability

    def get_id(self): return self._id
    def get_name(self): return self._name
    def get_availability(self): return self._availability
    def __str__(self): return self._name


class Room:
    def __init__(self, number, capacity):
        self._number = number
        self._capacity = capacity

    def get_number(self): return self._number
    def get_capacity(self): return self._capacity


class MeetingTime:
    def __init__(self, id, name, time):
        self._id = id
        self._name = name
        self._time = time

    def get_id(self): return self._id
    def get_name(self): return self._name
    def get_time(self): return self._time


class Department:
    def __init__(self, name, courses):
        self._name = name
        self._courses = courses

    def get_name(self): return self._name
    def get_courses(self): return self._courses


class Class:
    def __init__(self, id, dept, course):
        self._id = id
        self._dept = dept
        self._course = course
        self._instructor = None
        self._meetingTime = None
        self._room = None

    def get_id(self): return self._id
    def get_dept(self): return self._dept
    def get_course(self): return self._course
    def get_instructor(self): return self._instructor
    def get_meetingTime(self): return self._meetingTime
    def get_room(self): return self._room
    def set_instructor(self, instructor): self._instructor = instructor
    def set_meetingTime(self, meetingTime): self._meetingTime = meetingTime
    def set_room(self, room): self._room = room

    def __str__(self):
        return str(self._dept.get_name()) + "," + str(self._course.get_number()) + "," + \
            str(self._room.get_number()) + "," + str(self._instructor.get_id()
                                                     ) + "," + str(self._meetingTime.get_name())


class Conflict:
    class ConflictType(Enum):
        INSTRUCTOR_BOOKING = 1
        ROOM_BOOKING = 2
        จำนวนนักศึกษาเกิน = 3
        INSTRUCTOR_AVAILABILITY = 4
        ปีมีวันซ้ำกัน = 5
        ไม่ไขว้ = 6
        ยกเว้นการสอนของชั้นปี = 7
        สอนขนาน = 8

    def __init__(self, conflictType, conflictBetweenClasses):
        self._conflictType = conflictType
        self._conflictBetweenClasses = conflictBetweenClasses

    def get_conflictType(self): return self._conflictType
    def get_conflictBetweenClasses(self): return self._conflictBetweenClasses
    def __str__(self): return str(self._conflictType)+" " + \
        str(" and ".join(map(str, self._conflictBetweenClasses)))


class DisplayMgr:

    @staticmethod
    def display_generation(population):
        table1 = prettytable.PrettyTable(
            ['schedule #', 'fitness', '# of conflicts'])
        schedules = population.get_schedules()
        for i in range(0, len(schedules)):
            table1.add_row([str(i+1), round(schedules[i].get_fitness(), 3),
                           len(schedules[i].get_conflicts())])
        print(table1)

    @staticmethod
    def display_schedule_as_table(schedule):
        classes = schedule.get_classes()
        table = prettytable.PrettyTable(
            ['Class #', 'Dept', 'Course (number, max # of students)', 'Room (Capacity)', 'Instructor (Id)',  'Meeting Time (Id)'])
        for i in range(0, len(classes)):
            table.add_row([str(i+1), classes[i].get_dept().get_name(), classes[i].get_course().get_name() + " (" +
                           classes[i].get_course().get_number() + ", " +
                           str(classes[i].get_course(
                           ).get_maxNumbOfStudents()) + ")",
                           classes[i].get_room().get_number(
            ) + " (" + str(classes[i].get_room().get_capacity()) + ")",
                classes[i].get_instructor().get_name(
            ) + " (" + str(classes[i].get_instructor().get_id()) + ")",
                classes[i].get_meetingTime().get_time() + " (" + str(classes[i].get_meetingTime().get_id()) + ")"])
        print(table)

    @staticmethod
    def display_schedule_conflicts(schedule):
        conflictsTable = prettytable.PrettyTable(
            ['conflict type', 'conflict between classes'])
        conflicts = schedule.get_conflicts()
        for i in range(0, len(conflicts)):
            conflictsTable.add_row([str(conflicts[i].get_conflictType()),
                                    str("  and  ".join(map(str, conflicts[i].get_conflictBetweenClasses())))])
        if (len(conflicts) > 0):
            print(conflictsTable)


def find_fittest_schedule(verboseFlag):

    generationNumber = 0
    if (verboseFlag):
        print("> Generation # "+str(generationNumber))
    population = Population(POPULATION_SIZE)
    population.get_schedules().sort(key=lambda x: x.get_fitness(), reverse=True)

    # for i in range(0, 9):#
    #     print("population ", i+1)#
    #     DisplayMgr.display_schedule_as_table(population.get_schedules()[i])#

    # print()#
    # print("เรียงค่า fitness :")#
    # DisplayMgr.display_generation(population)#
    geneticAlgorithm = GeneticAlgorithm()
    run = 0
    while (population.get_schedules()[0].get_fitness() != 1.0):
        generationNumber += 1
        if (verboseFlag):
            print("\n> Generation # " + str(generationNumber))
        population = geneticAlgorithm.evolve(population)

        # print("-"*300)#
        # print("ยังไม่ได้เรียง :")#
        # population.get_schedules()#
        # DisplayMgr.display_generation(population)#

        # เอา schedule ที่ได้มา เอาไปคิดค่า fitness แล้วเรียงจากมากไปน้อย
        population.get_schedules().sort(key=lambda x: x.get_fitness(), reverse=True)
        # if (verboseFlag):
        #     print("ผลลัพท์ที่เรียง fitness :")
        #     DisplayMgr.display_generation(population)

        # for i in range(0,9):
        #     print("schedule ", i+1)
        # DisplayMgr.display_schedule_as_table(population.get_schedules()[0])
        DisplayMgr.display_schedule_conflicts(
            population.get_schedules()[0])

        # run+=1
        # if run == 2:
        # break
    print("> solution found after " + str(generationNumber) + " generations")
    # population_quantity = []
    # for x in range(number_find_fittest_schedule):
    #     population_quantity.append(population.get_schedules()[x])
    # return population_quantity
    return [population.get_schedules()[0], generationNumber]


dbMgr = DBMgr()


@app.route('/')
def timetabling():
    return render_template("process.html")


@app.route('/find_fittest_schedule', methods=['POST'])
def displayFitness():
    global schedule
    global number_find_fittest_schedule
    if request.method == 'POST':
        number_find_fittest_schedule = int(request.form.get('number'))
        temp =  [find_fittest_schedule(
            True) for y in range(number_find_fittest_schedule)]
        generationNumbers = [x[1] for x in temp]
        schedule = [x[0] for x in temp]
        classes = [x.get_classes() for x in schedule]

        rooms = [[] for i in range(len(classes))]
        numberOfRooms = [0 for i in range(len(classes))]
        for x in range(0, len(classes)):
            for i in range(0, len(classes[x])):
                rooms[x].append(classes[x][i].get_room().get_number())
            numberOfRooms[x] = len(set(rooms[x]))

        return render_template("selectResult.html", classes=classes, number_find_fittest_schedule=number_find_fittest_schedule, numberOfRooms=numberOfRooms, generationNumbers=generationNumbers)


@ app.route('/select_Number', methods=['POST'])
def select_Number():
    global schedule
    if request.method == 'POST':
        number = int(request.form['number'])
        schedule = schedule[number]
        return redirect('/display_schedule_as_table')


@ app.route('/display_schedule_as_table', methods=['GET'])
def displayTable():
    if request.method == 'GET':
        classes = schedule.get_classes()
        return render_template("timetabling.html", classes=classes)


@ app.route('/export_table')
def exportTable():
    # Get classes from schedule
    classes = schedule.get_classes()
    # Create a new workbook
    workbook = Workbook()
    del workbook['Sheet']
    sheet1 = workbook.create_sheet("วิทยาการคอมพิวเตอร์ ชั้นปีที่ 1")
    sheet2 = workbook.create_sheet("วิทยาการคอมพิวเตอร์ ชั้นปีที่ 2")
    sheet3 = workbook.create_sheet("วิทยาการคอมพิวเตอร์ ชั้นปีที่ 3")
    sheet4 = workbook.create_sheet("วิทยาการคอมพิวเตอร์ ชั้นปีที่ 4")
    sheet1.column_dimensions[get_column_letter(1)].width = 10
    sheet1.column_dimensions[get_column_letter(2)].width = 70
    sheet1.column_dimensions[get_column_letter(3)].width = 70
    sheet2.column_dimensions[get_column_letter(1)].width = 10
    sheet2.column_dimensions[get_column_letter(2)].width = 70
    sheet2.column_dimensions[get_column_letter(3)].width = 70
    sheet3.column_dimensions[get_column_letter(1)].width = 10
    sheet3.column_dimensions[get_column_letter(2)].width = 70
    sheet3.column_dimensions[get_column_letter(3)].width = 70
    sheet4.column_dimensions[get_column_letter(1)].width = 10
    sheet4.column_dimensions[get_column_letter(2)].width = 70
    sheet4.column_dimensions[get_column_letter(3)].width = 70

    allsheet = [sheet1, sheet2, sheet3, sheet4]
    dept = ['วิทยาการคอมพิวเตอร์ ชั้นปีที่ 1', 'วิทยาการคอมพิวเตอร์ ชั้นปีที่ 2',
            'วิทยาการคอมพิวเตอร์ ชั้นปีที่ 3', 'วิทยาการคอมพิวเตอร์ ชั้นปีที่ 4']

    # Add timetable data
    for i in range(4):
        check = [x for x in classes if x.get_dept().get_name() ==
                 dept[i]]
        day = {"จ.": {"09:00-12:00": "", "13:00-16:00": ""},
               "อ.": {"09:00-12:00": "", "13:00-16:00": ""},
               "พ.": {"09:00-12:00": "", "13:00-16:00": ""},
               "พฤ.": {"09:00-12:00": "", "13:00-16:00": ""},
               "ศ.": {"09:00-12:00": "", "13:00-16:00": ""}}
        for cls in check:
            if cls.get_course().get_number()[-1] == "A":
                xxx = " (กลุ่ม 2)"
            else:
                xxx = " (กลุ่ม 1)"

            day[cls.get_meetingTime().get_time().split(" ")[0]][cls.get_meetingTime(
            ).get_time().split(" ")[1]] += cls.get_course().get_name() + " " + xxx + " " + cls.get_instructor().get_name() + "\n"

        headers = ['วัน-เวลา', '09:00-12:00', '13:00-16:00']
        allsheet[i].append(headers)

        for day, timeslots in day.items():
            row = [day, timeslots.get('09:00-12:00', ''),
                   timeslots.get('13:00-16:00', '')]
            allsheet[i].append(row)

    # Save the workbook to a BytesIO object
    file_data = BytesIO()
    workbook.save(file_data)
    file_data.seek(0)

    # Create a response with the file data
    response = make_response(file_data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=timetable.xlsx'
    response.mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@ app.route('/display_schedule_meetingTimes', methods=['GET'])
def displayMeetingTimes():
    if request.method == 'GET':
        meetingTimes = dbMgr.get_meetingTimes()
        classes = schedule.get_classes()

        # Prepare the data for the template context
        meetingTimesTable = []
        for meetingTime in meetingTimes:
            matchingClasses_course = [
                str(cls.get_course().get_name()) for cls in classes if cls.get_meetingTime() == meetingTime]
            matchingClasses_dept = [
                str(cls.get_dept().get_name()) for cls in classes if cls.get_meetingTime() == meetingTime]
            matchingClasses_room = [
                str(cls.get_room().get_number()) for cls in classes if cls.get_meetingTime() == meetingTime]
            matchingClasses_instructor = [
                str(cls.get_instructor().get_name()) for cls in classes if cls.get_meetingTime() == meetingTime
            ]
            meetingTimesTable.append({
                'name': meetingTime.get_name(),
                'meeting_time': meetingTime.get_time(),
                'dept': matchingClasses_dept,
                'course': matchingClasses_course,
                'room': matchingClasses_room,
                'instructor': matchingClasses_instructor
            })

        return render_template("timetabling.html",  meetingTimesTable=meetingTimesTable)


@ app.route('/export_timetable')
def exportTimetable():
    workbook = Workbook()
    del workbook['Sheet']
    meetingTimes = dbMgr.get_meetingTimes()
    classes = schedule.get_classes()
    for meetingTime in meetingTimes:
        sheet = workbook.create_sheet(meetingTime.get_name())
        sheet.column_dimensions[get_column_letter(1)].width = 30
        sheet.column_dimensions[get_column_letter(2)].width = 50
        sheet.column_dimensions[get_column_letter(3)].width = 10
        sheet.column_dimensions[get_column_letter(4)].width = 15

        headers = ['ชั้นปี', 'วิชา', 'ห้อง', 'อาจารย์']
        sheet.append(headers)
        matchingClasses_dept = [
            str(cls.get_dept().get_name()) for cls in classes if cls.get_meetingTime() == meetingTime]
        matchingClasses_course = [
            str(cls.get_course().get_name()) for cls in classes if cls.get_meetingTime() == meetingTime]
        matchingClasses_room = [
            str(cls.get_room().get_number()) for cls in classes if cls.get_meetingTime() == meetingTime]
        matchingClasses_instructor = [
            str(cls.get_instructor().get_name()) for cls in classes if cls.get_meetingTime() == meetingTime]

        for xx in range(len(matchingClasses_course)):
            row = [matchingClasses_dept[xx], matchingClasses_course[xx],
                   matchingClasses_room[xx], matchingClasses_instructor[xx]]
            sheet.append(row)

    file_data = BytesIO()
    workbook.save(file_data)
    file_data.seek(0)

    # Create a response with the file data
    response = make_response(file_data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=time_schedule.xlsx'
    response.mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@ app.route('/display_schedule_rooms', methods=['GET'])
def displayRooms():
    if request.method == 'GET':
        rooms = dbMgr.get_rooms()
        classes = schedule.get_classes()
        # Prepare the data for the template context
        scheduleRoomsTable = []
        for room in rooms:
            roomSchedule_dept = [
                str(cls.get_dept().get_name()) for cls in classes if cls.get_room() == room]
            roomSchedule_course = [
                str(cls.get_course().get_name()) for cls in classes if cls.get_room() == room]
            roomSchedule_instructor = [
                str(cls.get_instructor().get_name()) for cls in classes if cls.get_room() == room]
            roomSchedule_meetingtime = [
                str(cls.get_meetingTime().get_time()) for cls in classes if cls.get_room() == room]

            scheduleRoomsTable.append({
                'room': str(room.get_number()),
                'dept': roomSchedule_dept,
                'course': roomSchedule_course,
                'coursenumber': roomSchedule_course,
                'instructor': roomSchedule_instructor,
                'meetingtime': roomSchedule_meetingtime

            })
        return render_template("timetabling.html", scheduleRoomsTable=scheduleRoomsTable)


@ app.route('/export_room')
def exportRoom():
    workbook = Workbook()
    del workbook['Sheet']
    rooms = dbMgr.get_rooms()
    classes = schedule.get_classes()
    for room in rooms:
        sheet = workbook.create_sheet(room.get_number())
        headers = ['วัน-เวลา', '09:00-12:00', '13:00-16:00']
        sheet.append(headers)
        sheet.column_dimensions[get_column_letter(1)].width = 10
        sheet.column_dimensions[get_column_letter(2)].width = 50
        sheet.column_dimensions[get_column_letter(3)].width = 50
        check = [x for x in classes if x.get_room() == room]
        day = {"จ.": {"09:00-12:00": "", "13:00-16:00": ""},
               "อ.": {"09:00-12:00": "", "13:00-16:00": ""},
               "พ.": {"09:00-12:00": "", "13:00-16:00": ""},
               "พฤ.": {"09:00-12:00": "", "13:00-16:00": ""},
               "ศ.": {"09:00-12:00": "", "13:00-16:00": ""}}
        for cls in check:
            day[cls.get_meetingTime().get_time().split(" ")[0]][cls.get_meetingTime(
            ).get_time().split(" ")[1]] += cls.get_course().get_name() + "\n"

        for day, timeslots in day.items():
            row = [day, timeslots.get('09:00-12:00', ''),
                   timeslots.get('13:00-16:00', '')]
            sheet.append(row)

    file_data = BytesIO()
    workbook.save(file_data)
    file_data.seek(0)

    # Create a response with the file data
    response = make_response(file_data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=room_schedule.xlsx'
    response.mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@ app.route('/display_schedule_instructors', methods=['GET'])
def displayInstructors():
    if request.method == 'GET':
        instructors = dbMgr.get_instructors()
        classes = schedule.get_classes()
        # Prepare the data for the template context
        instructorsTable = []

        for instructor in instructors:
            instructor_dept = [
                str(cls.get_dept().get_name()) for cls in classes if cls.get_instructor() == instructor]
            instructor_course = [
                str(cls.get_course().get_name()) for cls in classes if cls.get_instructor() == instructor]
            instructor_room = [
                str(cls.get_room().get_number()) for cls in classes if cls.get_instructor() == instructor]
            instructor_meetingtime = [
                str(cls.get_meetingTime().get_time()) for cls in classes if cls.get_instructor() == instructor]

            instructorsTable.append({
                'id': instructor.get_id(),
                'instructor': instructor.get_name(),
                'dept': instructor_dept,
                'course': instructor_course,
                'room': instructor_room,
                'meetingtime': instructor_meetingtime,
            })
        return render_template("timetabling.html", instructorsTable=instructorsTable)


@ app.route('/export_schedule_instructors')
def exportInstructors():
    workbook = Workbook()
    del workbook['Sheet']
    instructors = dbMgr.get_instructors()
    classes = schedule.get_classes()
    for instructor in instructors:
        sheet = workbook.create_sheet(instructor.get_name())
        headers = ['วัน-เวลา', '09:00-12:00', '13:00-16:00']
        sheet.append(headers)
        sheet.column_dimensions[get_column_letter(1)].width = 10
        sheet.column_dimensions[get_column_letter(2)].width = 50
        sheet.column_dimensions[get_column_letter(3)].width = 50
        check = [x for x in classes if x.get_instructor() == instructor]
        day = {"จ.": {"09:00-12:00": "", "13:00-16:00": ""},
               "อ.": {"09:00-12:00": "", "13:00-16:00": ""},
               "พ.": {"09:00-12:00": "", "13:00-16:00": ""},
               "พฤ.": {"09:00-12:00": "", "13:00-16:00": ""},
               "ศ.": {"09:00-12:00": "", "13:00-16:00": ""}}
        for cls in check:
            day[cls.get_meetingTime().get_time().split(" ")[0]][cls.get_meetingTime(
            ).get_time().split(" ")[1]] += cls.get_course().get_name() + "\n"

        for day, timeslots in day.items():
            row = [day, timeslots.get('09:00-12:00', ''),
                   timeslots.get('13:00-16:00', '')]
            sheet.append(row)

    file_data = BytesIO()
    workbook.save(file_data)
    file_data.seek(0)

    response = make_response(file_data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=instructor.xlsx'
    response.mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@ app.route('/dbteacher', methods=['GET', 'POST'])
def dbteacher():
    if request.method == 'POST':
        action = request.form['action']
        name = request.form['name']

        if action == "ADD":
            number = generate_number()
            dbMgr._c.execute(
                'INSERT INTO instructor (number, name) VALUES (%s, %s)', (number, name))
        elif action == "UPDATE":
            id = request.form['id']
            dbMgr._c.execute(
                """UPDATE instructor SET name=%s WHERE number=%s""", (name, id))

        dbMgr._conn.commit()
        return redirect('/dbteacher')

    if request.method == 'GET':
        dbMgr._c.execute("SELECT * FROM instructor")
        instructors = dbMgr._c.fetchall()
        return render_template("dbteacher.html", instructors=instructors)


def generate_number():
    dbMgr._c.execute("SELECT MAX(number) FROM instructor")
    result = dbMgr._c.fetchone()
    if result[0] is not None:
        last_number = result[0]
        prefix = last_number[0]
        last_numeric_part = int(last_number[1:])
        number = last_numeric_part + 1
        while True:
            new_number = f"{prefix}{number}"
            dbMgr._c.execute(
                "SELECT COUNT(*) FROM instructor WHERE number = %s", (new_number,))
            count = dbMgr._c.fetchone()[0]
            if count == 0:
                break
            number += 1
    else:
        new_number = "I1"
    return new_number


@ app.route('/dbteacher/delete/<id>')
def deleteTea(id):
    sql = "DELETE FROM instructor WHERE number = %s"
    val = (id,)
    dbMgr._c.execute(sql, val)
    dbMgr._conn.commit()
    return redirect('/dbteacher')


@ app.route('/dbcourse', methods=['GET', 'POST'])
def dbcourse():
    if request.method == 'POST':
        if request.form['action'] == 'ADD':
            number = request.form['number']
            name = request.form['name']
            max_numb_of_students = 0
            dbMgr._c.execute(
                'INSERT INTO course (number, name, max_numb_of_students) VALUES (%s, %s, %s)',
                (number, name, max_numb_of_students))
            dbMgr._conn.commit()
            return redirect('/dbcourse')

        elif request.form['action'] == 'UPDATE':
            number = request.form['number']
            name = request.form['name']
            dbMgr._c.execute(
                'UPDATE course SET name=%s WHERE number=%s', (name, number))
            dbMgr._conn.commit()
            return redirect('/dbcourse')

    if request.method == 'GET':
        dbMgr._c.execute("SELECT number, name FROM course")
        courses = dbMgr._c.fetchall()
        return render_template("dbcourse.html", courses=courses)
    return "Invalid request"


@ app.route('/dbcourse/delete/<id>')
def deleteCourse(id):
    sql = "DELETE FROM course WHERE number = %s"
    val = (id,)
    dbMgr._c.execute(sql, val)
    dbMgr._conn.commit()
    return redirect('/dbcourse')


@ app.route('/dbroom', methods=['GET', 'POST'])
def dbroom():
    if request.method == 'POST':
        action = request.form['action']
        number = request.form['number']
        capacity = request.form['capacity']

        if action == "ADD":
            dbMgr._c.execute(
                'INSERT INTO room (number, capacity) VALUES (%s, %s)', (number, capacity))
        elif action == "UPDATE":
            id = request.form['id']
            dbMgr._c.execute(
                """UPDATE room SET capacity=%s WHERE number=%s""", (capacity, id))

        dbMgr._conn.commit()
        return redirect('/dbroom')

    if request.method == 'GET':
        dbMgr._c.execute("SELECT * FROM room")
        room = dbMgr._c.fetchall()
        return render_template("dbroom.html", room=room)


@ app.route('/dbroom/delete/<id>')
def deleteRoom(id):
    sql = "DELETE FROM room WHERE number = %s"
    val = (id,)
    dbMgr._c.execute(sql, val)
    dbMgr._conn.commit()
    return redirect('/dbroom')


@ app.route('/matchCourse', methods=['GET', 'POST'])
def matchCourse():
    if request.method == 'POST':
        if request.form['action'] == 'ADD':
            course = request.form['course']
            course_values = course.split()
            number = course_values[0]
            name = course_values[1]
            dept = request.form['dept']
            instructor = request.form['instructor']
            instructor2 = request.form['instructor2']
            max_numb_of_students = request.form['max_numb_of_students']
            max_numb_of_students_2 = request.form['max_numb_of_students_2']
            checkbox_value = request.form.get('checkbox')
            selectbox_value = request.form.get('checkbox2')
            if checkbox_value:
                dbMgr._c.execute(
                    """UPDATE course SET name=%s , max_numb_of_students=%s WHERE number=%s""", (name, max_numb_of_students, number))
                dbMgr._conn.commit()
                dbMgr._c.execute(
                    'INSERT INTO course (number, name, max_numb_of_students) VALUES (%s, %s, %s)', (f'{number}A', name, max_numb_of_students_2))
                dbMgr._conn.commit()
                dbMgr._c.execute(
                    "INSERT INTO dept_course (name, course_numb, sec) VALUES (%s , %s, %s)", (dept, number, "1"))
                dbMgr._conn.commit()
                dbMgr._c.execute(
                    "INSERT INTO dept_course (name, course_numb, sec) VALUES (%s , %s, %s)", (dept, f'{number}A', "2"))
                dbMgr._conn.commit()
                dbMgr._c.execute(
                    "INSERT INTO course_instructor (course_number, instructor_number) VALUES (%s , %s)", (number, instructor))
                dbMgr._conn.commit()
                dbMgr._c.execute(
                    "INSERT INTO course_instructor (course_number, instructor_number) VALUES (%s , %s)", (f'{number}A', instructor))
                dbMgr._conn.commit()

            elif selectbox_value:
                dbMgr._c.execute(
                    """UPDATE course SET name=%s , max_numb_of_students=%s WHERE number=%s""", (name, max_numb_of_students, number))
                dbMgr._conn.commit()
                dbMgr._c.execute(
                    'INSERT INTO course (number, name, max_numb_of_students) VALUES (%s, %s, %s)', (f'{number}A', name, max_numb_of_students_2))
                dbMgr._conn.commit()
                dbMgr._c.execute(
                    "INSERT INTO dept_course (name, course_numb, sec) VALUES (%s , %s, %s)", (dept, number, "1"))
                dbMgr._conn.commit()
                dbMgr._c.execute(
                    "INSERT INTO dept_course (name, course_numb, sec) VALUES (%s , %s, %s)", (dept, f'{number}A', "2"))
                dbMgr._conn.commit()
                dbMgr._c.execute(
                    "INSERT INTO course_instructor (course_number, instructor_number) VALUES (%s , %s)", (number, instructor))
                dbMgr._conn.commit()
                dbMgr._c.execute(
                    "INSERT INTO course_instructor (course_number, instructor_number) VALUES (%s , %s)", (f'{number}A', instructor2))
                dbMgr._conn.commit()
                dbMgr._c.execute(
                    "INSERT INTO parallel (name1_parallel, name2_parallel) VALUES (%s , %s)", (number, f'{number}A'))
                dbMgr._conn.commit()

            else:
                dbMgr._c.execute(
                    """UPDATE course SET name=%s , max_numb_of_students=%s WHERE number=%s""", (name, max_numb_of_students, number))
                dbMgr._conn.commit()
                dbMgr._c.execute(
                    "INSERT INTO dept_course (name, course_numb, sec) VALUES (%s , %s, %s)", (dept, number, "1"))
                dbMgr._conn.commit()
                dbMgr._c.execute(
                    "INSERT INTO course_instructor (course_number, instructor_number) VALUES (%s , %s)", (number, instructor))
                dbMgr._conn.commit()

        elif request.form['action'] == 'UPDATE':
            course = request.form['course']
            course_values = course.split()
            number = course_values[0]
            name = course_values[1]
            dept = request.form['dept']
            instructor = request.form['instructor']
            max_numb_of_students = request.form['max_numb_of_students']
            dbMgr._c.execute(
                """UPDATE course SET max_numb_of_students=%s WHERE number=%s""", (max_numb_of_students, number))
            dbMgr._conn.commit()
            dbMgr._c.execute(
                """UPDATE dept_course SET name=%s WHERE course_numb=%s """, (dept, number))
            dbMgr._conn.commit()
            dbMgr._c.execute(
                """UPDATE course_instructor SET instructor_number=%s WHERE course_number=%s""", (instructor, number))
            dbMgr._conn.commit()
            return redirect('/matchCourse')
        return redirect('/matchCourse')

    if request.method == 'GET':
        dbMgr._c.execute("SELECT * FROM course")
        course = dbMgr._c.fetchall()
        dbMgr._c.execute("SELECT * FROM dept")
        dept = dbMgr._c.fetchall()
        dbMgr._c.execute("SELECT * FROM instructor")
        instructor = dbMgr._c.fetchall()
        dbMgr._c.execute(
            "SELECT dept_course.name, course.name, dept_course.sec, instructor.name, course.max_numb_of_students, dept_course.course_numb, instructor.number FROM dept_course JOIN course ON dept_course.course_numb = course.number JOIN course_instructor ON dept_course.course_numb = course_instructor.course_number JOIN instructor ON course_instructor.instructor_number = instructor.number WHERE dept_course.name = 'วิทยาการคอมพิวเตอร์ ชั้นปีที่ 1' ORDER BY dept_course.name;")
        dept_course1 = dbMgr._c.fetchall()
        dbMgr._c.execute(
            "SELECT dept_course.name, course.name, dept_course.sec, instructor.name, course.max_numb_of_students, dept_course.course_numb, instructor.number FROM dept_course JOIN course ON dept_course.course_numb = course.number JOIN course_instructor ON dept_course.course_numb = course_instructor.course_number JOIN instructor ON course_instructor.instructor_number = instructor.number WHERE dept_course.name = 'วิทยาการคอมพิวเตอร์ ชั้นปีที่ 2' ORDER BY dept_course.name;")
        dept_course2 = dbMgr._c.fetchall()
        dbMgr._c.execute(
            "SELECT dept_course.name, course.name, dept_course.sec, instructor.name, course.max_numb_of_students, dept_course.course_numb, instructor.number FROM dept_course JOIN course ON dept_course.course_numb = course.number JOIN course_instructor ON dept_course.course_numb = course_instructor.course_number JOIN instructor ON course_instructor.instructor_number = instructor.number WHERE dept_course.name = 'วิทยาการคอมพิวเตอร์ ชั้นปีที่ 3' ORDER BY dept_course.name;")
        dept_course3 = dbMgr._c.fetchall()
        dbMgr._c.execute(
            "SELECT dept_course.name, course.name, dept_course.sec, instructor.name, course.max_numb_of_students, dept_course.course_numb, instructor.number FROM dept_course JOIN course ON dept_course.course_numb = course.number JOIN course_instructor ON dept_course.course_numb = course_instructor.course_number JOIN instructor ON course_instructor.instructor_number = instructor.number WHERE dept_course.name = 'วิทยาการคอมพิวเตอร์ ชั้นปีที่ 4' ORDER BY dept_course.name;")
        dept_course4 = dbMgr._c.fetchall()

        return render_template("matchCourse.html", course=course, dept=dept, instructor=instructor, dept_course1=dept_course1, dept_course2=dept_course2, dept_course3=dept_course3, dept_course4=dept_course4)


@ app.route('/matchCourse/delete/<id>', methods=['GET', 'POST'])
def deleteMatchCourse(id):
    if request.method == 'POST':
        sql_dept_course = "DELETE FROM dept_course WHERE course_numb = %s"
        val_dept_course = (id,)
        dbMgr._c.execute(sql_dept_course, val_dept_course)
        dbMgr._conn.commit()

        sql_course_instructor = "DELETE FROM course_instructor WHERE course_number = %s"
        val_course_instructor = (id,)
        dbMgr._c.execute(sql_course_instructor, val_course_instructor)
        dbMgr._conn.commit()

        sql_parallel = "DELETE FROM parallel WHERE name1_parallel = %s OR name2_parallel = %s;"
        val_parallel = (id, id)
        dbMgr._c.execute(sql_parallel, val_parallel)
        dbMgr._conn.commit()

        if len(id) == 9:
            sql_course = "DELETE FROM course WHERE number = %s;"
            val_course = (id,)
            dbMgr._c.execute(sql_course, val_course)
            dbMgr._conn.commit()

        elif len(id) == 8:
            sql_course_update = "UPDATE course SET max_numb_of_students = 0 WHERE number = %s;"
            val_course_update = (id,)
            dbMgr._c.execute(sql_course_update, val_course_update)
            dbMgr._conn.commit()

        return redirect('/matchCourse')


@ app.route('/multiSec', methods=['GET', 'POST'])
def multiSec():
    if request.method == 'POST':
        course1 = request.form['course1']
        course2 = request.form['course2']
        dbMgr._c.execute(
            "INSERT INTO multi_sec (name1, name2) VALUES (%s , %s)", (course1, f'{course2}A'))
        dbMgr._c.execute(
            "INSERT INTO multi_sec (name1, name2) VALUES (%s , %s)", (f'{course1}A', course2))
        dbMgr._conn.commit()
        return redirect('/multiSec')

    if request.method == 'GET':
        dbMgr._c.execute("SELECT * FROM dept")
        dept = dbMgr._c.fetchall()
        dbMgr._c.execute(
            "SELECT number, name FROM course WHERE SUBSTRING(number, 1, 8) = SUBSTRING(number, 1, 8) GROUP BY SUBSTRING(number, 1, 8) HAVING COUNT(*) > 1;")
        course = dbMgr._c.fetchall()
        dbMgr._c.execute("SELECT course1.name , course2.name , multi_sec.name1 FROM multi_sec JOIN course AS course1 ON multi_sec.name1 = course1.number JOIN course AS course2 ON multi_sec.name2 = course2.number;")
        course_multi = dbMgr._c.fetchall()
        return render_template("multiSec.html", dept=dept, course=course, course_multi=course_multi)


@ app.route('/multiSec/delete/<id>', methods=['GET', 'POST'])
def deleteMultiSec(id):
    if request.method == 'POST':
        sql = "DELETE FROM multi_sec WHERE name1 = %s"
        val = (id,)
        dbMgr._c.execute(sql, val)
        dbMgr._conn.commit()
        return redirect('/multiSec')


@ app.route('/exp', methods=['GET', 'POST'])
def exp():
    nlist = []
    if request.method == 'POST':
        instructor_id = request.form['instructor_id']
        mon1 = request.form.get('1', None)
        mon2 = request.form.get('2', None)
        tue1 = request.form.get('3', None)
        tue2 = request.form.get('4', None)
        wed1 = request.form.get('5', None)
        wed2 = request.form.get('6', None)
        thu1 = request.form.get('7', None)
        thu2 = request.form.get('8', None)
        fri1 = request.form.get('9', None)
        fri2 = request.form.get('10', None)
        if mon1 is not None:
            nlist.append("1")
        if mon2 is not None:
            nlist.append("2")
        if tue1 is not None:
            nlist.append("3")
        if tue2 is not None:
            nlist.append("4")
        if wed1 is not None:
            nlist.append("5")
        if wed2 is not None:
            nlist.append("6")
        if thu1 is not None:
            nlist.append("7")
        if thu2 is not None:
            nlist.append("8")
        if fri1 is not None:
            nlist.append("9")
        if fri2 is not None:
            nlist.append("10")

        dbMgr._c.execute("SELECT * FROM course_instructor")
        instructor = dbMgr._c.fetchall()
        for i in nlist:
            dbMgr._c.execute(
                "INSERT INTO instructor_availability (instructor_id, meeting_time_id) VALUES (%s , %s)", (instructor_id, i))
            dbMgr._conn.commit()
        return redirect('/exp')

    if request.method == 'GET':
        dbMgr._c.execute("SELECT * FROM instructor")
        instructor = dbMgr._c.fetchall()
        dbMgr._c.execute("SELECT instructor.number, instructor.name, GROUP_CONCAT(meeting_time.name) AS meeting_times FROM instructor JOIN instructor_availability ON instructor.number = instructor_availability.instructor_id JOIN meeting_time ON instructor_availability.meeting_time_id = meeting_time.id GROUP BY instructor.number, instructor.name; ")
        instructor_availability = dbMgr._c.fetchall()
        return render_template("exp.html", instructor=instructor, instructor_availability=instructor_availability)


@ app.route('/exp/delete/<id>')
def deleteInsAvailability(id):
    sql = "DELETE FROM instructor_availability WHERE instructor_id = %s"
    val = (id,)
    dbMgr._c.execute(sql, val)
    dbMgr._conn.commit()
    return redirect('/exp')


@ app.route('/exp/dept', methods=['GET', 'POST'])
def expdept():
    nlist = []
    if request.method == 'POST':
        checkbox_value1 = request.form.get('checkbox1')
        checkbox_value2 = request.form.get('checkbox2')
        checkbox_value3 = request.form.get('checkbox3')
        checkbox_value4 = request.form.get('checkbox4')
        mon1 = request.form.get('1', None)
        mon2 = request.form.get('2', None)
        tue1 = request.form.get('3', None)
        tue2 = request.form.get('4', None)
        wed1 = request.form.get('5', None)
        wed2 = request.form.get('6', None)
        thu1 = request.form.get('7', None)
        thu2 = request.form.get('8', None)
        fri1 = request.form.get('9', None)
        fri2 = request.form.get('10', None)
        if mon1 is not None:
            nlist.append("1")
        if mon2 is not None:
            nlist.append("2")
        if tue1 is not None:
            nlist.append("3")
        if tue2 is not None:
            nlist.append("4")
        if wed1 is not None:
            nlist.append("5")
        if wed2 is not None:
            nlist.append("6")
        if thu1 is not None:
            nlist.append("7")
        if thu2 is not None:
            nlist.append("8")
        if fri1 is not None:
            nlist.append("9")
        if fri2 is not None:
            nlist.append("10")

        for i in nlist:
            if checkbox_value1:
                dbMgr._c.execute(
                    "SELECT * FROM exp_dept WHERE name_id = %s AND meeting_time = %s", ("วิทยาการคอมพิวเตอร์ ชั้นปีที่ 1", i))
                existing_record = dbMgr._c.fetchone()
                if not existing_record:
                    dbMgr._c.execute(
                        "INSERT INTO exp_dept (name_id, meeting_time) VALUES (%s , %s)", ("วิทยาการคอมพิวเตอร์ ชั้นปีที่ 1", i))
                    dbMgr._conn.commit()
            if checkbox_value2:
                dbMgr._c.execute(
                    "SELECT * FROM exp_dept WHERE name_id = %s AND meeting_time = %s", ("วิทยาการคอมพิวเตอร์ ชั้นปีที่ 2", i))
                existing_record = dbMgr._c.fetchone()
                if not existing_record:
                    dbMgr._c.execute(
                        "INSERT INTO exp_dept (name_id, meeting_time) VALUES (%s , %s)", ("วิทยาการคอมพิวเตอร์ ชั้นปีที่ 2", i))
                    dbMgr._conn.commit()
            if checkbox_value3:
                dbMgr._c.execute(
                    "SELECT * FROM exp_dept WHERE name_id = %s AND meeting_time = %s", ("วิทยาการคอมพิวเตอร์ ชั้นปีที่ 3", i))
                existing_record = dbMgr._c.fetchone()
                if not existing_record:
                    dbMgr._c.execute(
                        "INSERT INTO exp_dept (name_id, meeting_time) VALUES (%s , %s)", ("วิทยาการคอมพิวเตอร์ ชั้นปีที่ 3", i))
                    dbMgr._conn.commit()
            if checkbox_value4:
                dbMgr._c.execute(
                    "SELECT * FROM exp_dept WHERE name_id = %s AND meeting_time = %s", ("วิทยาการคอมพิวเตอร์ ชั้นปีที่ 4", i))
                existing_record = dbMgr._c.fetchone()
                if not existing_record:
                    dbMgr._c.execute(
                        "INSERT INTO exp_dept (name_id, meeting_time) VALUES (%s , %s)", ("วิทยาการคอมพิวเตอร์ ชั้นปีที่ 4", i))
                    dbMgr._conn.commit()
        return redirect('/exp/dept')

    if request.method == 'GET':
        dbMgr._c.execute("SELECT * FROM dept")
        dept = dbMgr._c.fetchall()
        dbMgr._c.execute(
            "SELECT exp_dept.name_id, GROUP_CONCAT(meeting_time.name) AS meeting_times FROM exp_dept JOIN meeting_time ON exp_dept.meeting_time = meeting_time.id GROUP BY exp_dept.name_id;")
        exp_dept = dbMgr._c.fetchall()
        return render_template("expdept.html", dept=dept, exp_dept=exp_dept)


@ app.route('/expdept/delete/<id>')
def deleteexpdept(id):
    sql = "DELETE FROM exp_dept WHERE name_id = %s"
    val = (id,)
    dbMgr._c.execute(sql, val)
    dbMgr._conn.commit()
    return redirect('/exp/dept')


@ app.route('/importcourse')
def importcourse():
    return render_template("importcourse.html")


@ app.route('/importcourse', methods=['POST'])
def uploadFilesCourse():
    if request.method == 'POST':
        uploaded_file = request.files['file']
        if uploaded_file.filename != '':
            file_path = os.path.join(
                app.config['UPLOAD_FOLDER'], uploaded_file.filename)
            uploaded_file.save(file_path)
            parseCSVcourse(file_path)
    return redirect(url_for('dbcourse'))


def parseCSVcourse(filePath):
    col_names = ['number', 'name', 'max_numb_of_students']
    csvData = pd.read_csv(filePath, names=col_names, header=None)
    for i, row in csvData.iterrows():
        sql = "INSERT INTO course (number,name,max_numb_of_students) VALUES (%s, %s, %s)"
        val = (str(row['number']).zfill(8),
               row['name'], row['max_numb_of_students'])
        dbMgr._c.execute(sql, val)
        dbMgr._conn.commit()


@ app.route('/importteacher')
def importteacher():
    return render_template("importteacher.html")


@ app.route('/importteacher', methods=['POST'])
def uploadFilesTeacher():
    if request.method == 'POST':
        uploaded_file = request.files['file']
        if uploaded_file.filename != '':
            file_path = os.path.join(
                app.config['UPLOAD_FOLDER'], uploaded_file.filename)
            uploaded_file.save(file_path)
            parseCSVteacher(file_path)
    return redirect(url_for('dbteacher'))


def parseCSVteacher(filePath):
    col_names = ['number', 'name']
    csvData = pd.read_csv(filePath, names=col_names, header=None)
    for i, row in csvData.iterrows():
        sql = "INSERT INTO instructor (number,name) VALUES (%s, %s)"
        val = (row['number'], row['name'])
        dbMgr._c.execute(sql, val)
        dbMgr._conn.commit()


@ app.route('/importroom')
def importroom():
    return render_template("importroom.html")


@ app.route('/importroom', methods=['POST'])
def uploadFilesroom():
    if request.method == 'POST':
        uploaded_file = request.files['file']
        if uploaded_file.filename != '':
            file_path = os.path.join(
                app.config['UPLOAD_FOLDER'], uploaded_file.filename)
            uploaded_file.save(file_path)
            parseCSVroom(file_path)
    return redirect(url_for('dbroom'))


def parseCSVroom(filePath):
    col_names = ['number', 'capacity']
    csvData = pd.read_csv(filePath, names=col_names, header=None)
    for i, row in csvData.iterrows():
        sql = "INSERT INTO room (number,capacity) VALUES (%s, %s)"
        val = (row['number'], row['capacity'])
        dbMgr._c.execute(sql, val)
        dbMgr._conn.commit()


if __name__ == "__main__":
    app.run(debug=True)
